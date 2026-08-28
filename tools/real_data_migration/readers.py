from __future__ import annotations
import csv
from pathlib import Path
from typing import Iterator
from .mapping import header_score, map_headers
from .normalize import text

class ReaderError(RuntimeError):
    pass


def _tabular_rows(path: Path) -> Iterator[tuple[str, list[list[object]]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        last_error = None
        for encoding in ("utf-8-sig", "cp1256"):
            try:
                with path.open("r", encoding=encoding, newline="") as f:
                    rows = [list(r) for r in csv.reader(f)]
                yield "CSV", rows
                return
            except UnicodeDecodeError as exc:
                last_error = exc
        raise ReaderError(f"Cannot decode CSV {path.name}: {last_error}")
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ReaderError("XLSX input requires openpyxl already present in the HRM build environment.") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                yield ws.title, [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
        return
    if suffix == ".xls":
        try:
            import xlrd  # type: ignore
        except ImportError as exc:
            raise ReaderError(
                "Legacy .xls input detected but xlrd is not available. "
                "Do not change the HRM dependency lock blindly; add/pin xlrd only after reviewing the baseline lock, "
                "or convert the private source workbook to .xlsx outside Git."
            ) from exc
        book = xlrd.open_workbook(path)
        for sheet in book.sheets():
            yield sheet.name, [sheet.row_values(i) for i in range(sheet.nrows)]
        return
    raise ReaderError(f"Unsupported input type: {path.suffix}")


def read_records(path: Path, min_header_score: int = 2) -> Iterator[tuple[str, int, dict[str, str]]]:
    for sheet_name, rows in _tabular_rows(path):
        if not rows:
            continue
        candidates = [(header_score(row), idx, row) for idx, row in enumerate(rows[:20])]
        score, header_idx, header_row = max(candidates, key=lambda x: x[0])
        if score < min_header_score:
            continue
        mapping = map_headers(header_row)
        for row_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
            record: dict[str, str] = {}
            for col_idx, field in mapping.items():
                value = row[col_idx] if col_idx < len(row) else None
                record[field] = text(value)
            if any(record.values()):
                yield sheet_name, row_idx, record
